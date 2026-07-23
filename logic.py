"""
Logic module for Product Feed Generator / Zecom Tracker processing.

- Maps output dynamically using header names from the Sample Output Sheet
  (with alias matching so differently-worded headers still map correctly).
- Region-aware: SG / MY / PH drive currency, and (when present) region-
  specific Price / Category / Size Chart tabs.
- User-selectable Price Tracker column (e.g. Selling Price, Promo Price,
  Marketplace Price) drives the itemAmount for every parent + child SKU.
- Category ID is resolved per Parent/Child SKU using the product Title +
  Gender against the Category sheet, picking the most specific match.
- Quantity (noOfItem) is always forced to 0 for every Parent + Child SKU.
- Sorts child variants sequentially by size order (Alpha/Numeric).
- Appends descriptions and care instructions into template attributes.
"""

import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Standard clothing size order for sorting
SIZE_ORDER = [
    "3XS", "XXXS", "XXS", "XS", "S", "S/M", "M", "M/L", "L", "L/XL",
    "XL", "XXL", "XXXL", "3XL", "4XL", "5XL", "6XL", "1-2Y", "2-3Y",
    "3-4Y", "4-5Y", "5-6Y", "6-7Y", "7-8Y", "8-9Y", "9-10Y", "10-11Y",
    "11-12Y", "12-13Y", "13-14Y", "14-15Y", "15-16Y", "6Y", "8Y", "10Y",
    "12Y", "14Y", "16Y", "18Y", "20Y", "OSFA", "ONE SIZE", "UA", "MINI",
    "KIDS", "ADULT", "YOUTH"
]

# Region -> default currency code
REGION_CURRENCY = {
    "SG": "SGD",
    "MY": "MYR",
    "PH": "PHP",
}

HEADINGS = [
    "SKU", "status", "errorDetails", "customSKU", "itemTitle", "itemDescription1",
    "itemDescription2", "itemDescription3", "noOfVariants", "variation1",
    "variation2", "variation3", "shortDescription", "salePrice", "saleStartDate",
    "saleEndDate", "itemAmount", "currencyCode", "noOfItem", "imageURI",
    "categoryID", "taxClass", "brand", "model", "warrantyType",
    "packageWeight(kg)", "packageHeight(cm)", "packageLength(cm)",
    "packageWidth(cm)", "packageContent", "itemSpecifics1", "itemSpecifics2",
    "itemSpecifics3", "itemSpecifics4", "itemSpecifics5", "itemSpecifics6",
    "itemSpecifics7", "itemSpecifics8", "itemSpecifics9", "itemSpecifics10",
    "itemSpecifics11", "itemSpecifics12", "itemSpecifics13", "itemSpecifics14",
    "itemSpecifics15", "itemSpecifics16", "itemSpecifics17", "itemSpecifics18",
    "itemSpecifics19", "itemSpecifics20", "itemSpecifics21", "itemSpecifics22",
    "itemSpecifics23", "itemSpecifics24", "itemSpecifics25", "templateAttribute1",
    "templateAttribute2", "templateAttribute3", "templateAttribute4",
    "templateAttribute5", "postAsNonVariant",
]


class ConversionError(Exception):
    """Raised for issues that prevent proper generation of the Output sheet."""
    pass


# ---------------------------------------------------------------------------
# Small generic helpers
# ---------------------------------------------------------------------------

def val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else v


def s(v):
    return "" if v is None else str(v)


def replace_first(text, old, new):
    return s(text).replace(old, new, 1)


def is_not_number(v):
    try:
        float(s(v).strip())
        return False
    except (ValueError, TypeError):
        return True


def normalize_header(h):
    """Lower-cases and strips everything but letters/digits, so headers like
    'Item Amount', 'item_amount' and 'itemAmount' all compare equal."""
    return re.sub(r"[^a-z0-9]", "", s(h).lower())


def parse_column_setting(ws, col_setting):
    """Resolves a column letter, 1-based index, or header name (row 1) on
    the given worksheet to a column index. Defaults to column 4 ('D')."""
    if not col_setting:
        return 4

    col_str = str(col_setting).strip()

    if col_str.isalpha():
        try:
            return column_index_from_string(col_str.upper())
        except ValueError:
            pass

    if col_str.isdigit():
        return int(col_str)

    for c in range(1, ws.max_column + 1):
        if str(val(ws, 1, c)).strip().lower() == col_str.lower():
            return c

    return 4


def find_header_col(ws, header_name):
    """Exact (case-insensitive, whitespace-trimmed) header match on row 1."""
    if not header_name:
        return None
    target = s(header_name).strip().lower()
    for c in range(1, ws.max_column + 1):
        if s(val(ws, 1, c)).strip().lower() == target:
            return c
    return None


def find_col_by_keywords(ws, keywords):
    """First column whose row-1 header contains any of the given keywords."""
    for c in range(1, ws.max_column + 1):
        h = s(val(ws, 1, c)).strip().lower()
        if any(k in h for k in keywords):
            return c
    return None


def get_sheet_headers(file_bytes, sheet_name=None):
    """Returns the row-1 header labels of a worksheet, given raw file bytes.
    If sheet_name is provided, tries to find that sheet (falling back to the
    active sheet); otherwise reads the active sheet. Used to power the
    Streamlit 'pick your price column' dropdown."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = None
    if sheet_name:
        ws = find_sheet(wb, sheet_name, required=False)
    if ws is None:
        ws = wb.active
    headers = []
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        if v not in (None, ""):
            headers.append(str(v).strip())
    return headers


# ---------------------------------------------------------------------------
# Output sheet construction (header-driven, alias aware)
# ---------------------------------------------------------------------------

# Canonical field -> itself, keyed by normalized text, used first.
_ALIAS_MAP = {normalize_header(h): h for h in HEADINGS}

# Common alternate wordings people use in a "Sample Output Sheet" that should
# still resolve to the same canonical field used throughout this module.
_EXTRA_ALIASES = {
    "quantity": "noOfItem",
    "qty": "noOfItem",
    "stock": "noOfItem",
    "stockqty": "noOfItem",
    "stockquantity": "noOfItem",
    "availableqty": "noOfItem",
    "price": "itemAmount",
    "sellingprice": "itemAmount",
    "regularprice": "itemAmount",
    "unitprice": "itemAmount",
    "amount": "itemAmount",
    "category": "categoryID",
    "categoryid": "categoryID",
    "categorycode": "categoryID",
    "image": "imageURI",
    "images": "imageURI",
    "imageurl": "imageURI",
    "imageuri": "imageURI",
    "title": "itemTitle",
    "producttitle": "itemTitle",
    "itemname": "itemTitle",
    "description": "itemDescription1",
    "productdescription": "itemDescription1",
    "shortdesc": "shortDescription",
    "variantsku": "customSKU",
    "childsku": "customSKU",
    "skuid": "customSKU",
    "parentsku": "SKU",
    "currency": "currencyCode",
    "brandname": "brand",
}
for _k, _v in _EXTRA_ALIASES.items():
    _ALIAS_MAP.setdefault(_k, _v)


def create_heading_for_target_sheet(main, custom_headings=None):
    """Writes the header row using the Sample Output Sheet's exact header
    text (so the output file matches it verbatim), while building a
    canonical-field -> column-index map so the rest of the code can keep
    writing values by canonical field name regardless of exact wording."""
    headings = custom_headings if custom_headings else HEADINGS
    col_map = {}
    for i, h in enumerate(headings):
        col_idx = i + 1
        main.set_value(1, col_idx, h)
        norm = normalize_header(h)
        canonical = _ALIAS_MAP.get(norm, s(h).strip())
        col_map.setdefault(canonical, col_idx)
        col_map.setdefault(s(h).strip(), col_idx)
    return col_map


def replace_spl_character(value):
    value = s(value)
    pairs = [
        ("â€œ", "\u201c"), ("â€", "\u201d"), ("â€˜", "\u2018"), ("â€™", "\u2019"),
        ("â€”", "\u2013"), ("â€“", "\u2014"), ("â€•", "-"), ("â€¦", "\u2026"),
        ("Ã˜", "\u00d8"), ("Ã‚Â®", "\u00ae"), ("Â³", "\u00b3"), ("Â®", "\u00ae"),
    ]
    for old, new in pairs:
        value = replace_first(value, old, new)
    return value


def remove_duplicates(title):
    parts = s(title).split(" ")
    result = []
    for p in parts:
        if p not in result:
            result.append(p)
    return " ".join(result)


class OutputSheet:
    def __init__(self):
        self.rows = {}
        self.max_col = 0
        self.max_row = 0

    def set_value(self, row, col, value):
        if col is None or col < 1:
            return
        self.rows.setdefault(row, {})[col] = value
        self.max_row = max(self.max_row, row)
        self.max_col = max(self.max_col, col)

    def to_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Output"
        for r in range(1, self.max_row + 1):
            for c in range(1, self.max_col + 1):
                v = self.rows.get(r, {}).get(c)
                if v is not None and v != "":
                    ws.cell(row=r, column=c, value=v)
        return wb


# ---------------------------------------------------------------------------
# Title building
# ---------------------------------------------------------------------------

def form_title(brand, new_regional_display_name, activity_group, article_type, gender, search_color_name, products_division):
    brand, new_regional_display_name, search_color_name = s(brand), s(new_regional_display_name), s(search_color_name)
    title = "[NEW] " + (replace_first(brand, "Licence", "PUMA") if "Licence" in brand else brand)

    if gender not in (None, "") and gender not in title and gender == "Unisex":
        title += " " + gender

    if new_regional_display_name not in title:
        if "Trainers" in new_regional_display_name or "Trainer" in new_regional_display_name:
            title += " " + replace_first(replace_first(new_regional_display_name, "Trainers", "Shoes"), "Trainer", "Shoes")
        elif "Sandals" in new_regional_display_name:
            title += " " + replace_first(new_regional_display_name, "Sandals", "Sports Sandals")
        elif "Slides" in new_regional_display_name:
            title += " " + replace_first(new_regional_display_name, "Slides", "Slides Slippers")
        else:
            title += " " + new_regional_display_name

    if products_division == "Footwear" and search_color_name not in title:
        title += " (" + search_color_name + ") "

    return title


def get_item_title(regional_display_name, brand, gender, activity_group, article_type, search_color_name, products_division):
    regional_display_name, search_color_name = s(regional_display_name), s(search_color_name)
    new_regional_display_name = replace_first(regional_display_name, "\u2019s", "'s\u2122") if "\u2019s" in regional_display_name else regional_display_name
    get_search_color_name = search_color_name.split(" - ")[1] if " - " in search_color_name else search_color_name

    if "Men" in regional_display_name or "Women" in regional_display_name:
        title = form_title(brand, new_regional_display_name, activity_group, article_type, "", get_search_color_name, products_division)
    else:
        title = form_title(brand, new_regional_display_name, activity_group, article_type, gender, get_search_color_name, products_division)
    return remove_duplicates(title)


# ---------------------------------------------------------------------------
# Size handling
# ---------------------------------------------------------------------------

def get_variation2_size(input_ws, i):
    """Extracts Size UK (Col 22) or calculates formatted variation string."""
    product_division = val(input_ws, i, 14)
    size_uk = val(input_ws, i, 22)
    size_fr = val(input_ws, i, 21)
    size_asia = val(input_ws, i, 23)
    size_us = val(input_ws, i, 20)

    if size_uk != "":
        return s(size_uk)

    variation = None
    if product_division in ("Footwear", "Accessories", "Socks"):
        if size_uk != "":
            variation = ("Int:" + s(size_uk)) if is_not_number(size_uk) else ("UK:" + s(size_uk))
        else:
            variation = ("Int:" + s(size_fr)) if is_not_number(size_fr) else ("US:" + s(size_fr))
    elif product_division == "Apparel":
        if size_uk != "":
            variation = ("Int:" + s(size_uk)) if is_not_number(size_uk) else ("UK:" + s(size_uk))
        elif size_us != "":
            variation = ("Int:" + s(size_us)) if is_not_number(size_us) else ("US:" + s(size_us))
        else:
            variation = ("Int:" + s(size_asia)) if is_not_number(size_asia) else ("ASIA:" + s(size_asia))

    return s(variation) if variation else ""


def size_sort_key(size_str):
    """Helper key generator for sorting variants by size."""
    clean_size = s(size_str).upper().replace("UK:", "").replace("US:", "").replace("INT:", "").replace("ASIA:", "").strip()
    if clean_size in SIZE_ORDER:
        return (0, SIZE_ORDER.index(clean_size))
    try:
        return (1, float(clean_size))
    except ValueError:
        return (2, clean_size)


def build_short_description(input_ws, idx):
    """Builds short description HTML bullet list from tracker metadata."""
    fields = [
        ("Brand", val(input_ws, idx, 3)),
        ("Color Name", val(input_ws, idx, 9)),  # Color Name (Col 9)
        ("Gender", val(input_ws, idx, 5)),
        ("Activity Group", val(input_ws, idx, 8)),
        ("Collection", val(input_ws, idx, 26)),
        ("Material", val(input_ws, idx, 27)),
        ("Upper Material", val(input_ws, idx, 29)),
        ("Mid Sole Material", val(input_ws, idx, 30)),
        ("Outer Sole Material", val(input_ws, idx, 31)),
        ("Style Number", val(input_ws, idx, 1)),
    ]
    items = [f"<li>{k} : {v}</li>" for k, v in fields if v not in ("", None, "Other")]
    return f"<ul>{''.join(items)}</ul>" if items else ""


# ---------------------------------------------------------------------------
# Price tracker handling (user-selectable price column)
# ---------------------------------------------------------------------------

def construct_amount_map(price_ws, price_column_name=None):
    """Builds sku -> {"price": ..., "sale_price": ...} using the user-chosen
    price column (e.g. 'Selling Price', 'Promo Price', 'Marketplace Price').
    Falls back to sensible defaults if headers can't be matched, so older
    trackers without a matching header still work."""
    sku_col = find_col_by_keywords(price_ws, ["sku"]) or 3

    price_col = find_header_col(price_ws, price_column_name)
    if price_col is None:
        price_col = find_col_by_keywords(price_ws, ["selling", "price", "rrp"]) or 4

    sale_col = find_col_by_keywords(price_ws, ["promo", "sale"])

    amount_map = {}
    for r in range(2, price_ws.max_row + 1):
        sku = val(price_ws, r, sku_col)
        if sku:
            amount_map[s(sku).strip()] = {
                "price": val(price_ws, r, price_col),
                "sale_price": val(price_ws, r, sale_col) if sale_col else "",
            }
    return amount_map


# ---------------------------------------------------------------------------
# Category mapping (by Title + Gender, most-specific match wins)
# ---------------------------------------------------------------------------

def construct_category_map(category_ws):
    """Builds a list of matching rules from the Category sheet.
    Expected layout: Col A = title keyword/phrase, Col B = Gender
    (optional), Col C = Category ID. Sheets with only two meaningful
    columns (keyword, category id) are also handled."""
    rules = []
    for r in range(2, category_ws.max_row + 1):
        keyword = s(val(category_ws, r, 1)).strip()
        if not keyword:
            continue
        col_b = val(category_ws, r, 2)
        col_c = val(category_ws, r, 3)

        if col_c in ("", None):
            # Only two populated columns: (keyword, category_id)
            gender = ""
            cat_id = col_b
        else:
            gender = s(col_b).strip()
            cat_id = col_c

        rules.append({"keyword": keyword, "gender": gender, "category_id": cat_id})
    return rules


def match_category(rules, title, gender):
    """Finds the most appropriate Category ID for a given item Title +
    Gender. Prefers the longest keyword match that also matches gender;
    falls back to the longest keyword match ignoring gender."""
    title_l = s(title).lower()
    gender_l = s(gender).strip().lower()

    best, best_len = None, -1
    for rule in rules:
        kw = rule["keyword"].lower()
        if kw and kw in title_l and rule["gender"] and rule["gender"].lower() == gender_l:
            if len(kw) > best_len:
                best, best_len = rule, len(kw)
    if best:
        return best["category_id"]

    best, best_len = None, -1
    for rule in rules:
        kw = rule["keyword"].lower()
        if kw and kw in title_l:
            if len(kw) > best_len:
                best, best_len = rule, len(kw)
    return best["category_id"] if best else ""


def get_template_attribute1(size_chart_ws):
    size_chart_map = {}
    for r in range(2, size_chart_ws.max_row + 1):
        key = val(size_chart_ws, r, 1)
        if key:
            size_chart_map[s(key).strip()] = [val(size_chart_ws, r, c) for c in range(1, 3)]
    return size_chart_map


def get_parent_key(input_ws, row_idx, products_division):
    if products_division in ["Apparel", "Accessories"]:
        return val(input_ws, row_idx, 1)
    elif products_division == "Footwear":
        return val(input_ws, row_idx, 9)
    return val(input_ws, row_idx, 1)


# ---------------------------------------------------------------------------
# Sheet lookup helpers (region aware)
# ---------------------------------------------------------------------------

def _normalize_sheet_name(name):
    return "".join(name.split()).lower()


def find_sheet(wb, expected_name, required=True):
    if expected_name in wb.sheetnames:
        return wb[expected_name]
    target = _normalize_sheet_name(expected_name)
    for name in wb.sheetnames:
        if _normalize_sheet_name(name) == target:
            return wb[name]
    if required:
        raise ConversionError(f"Could not find sheet '{expected_name}'. Available sheets: {wb.sheetnames}")
    return None


def find_regional_sheet(wb, base_name, region, required=False):
    """Looks for a region-specific tab first (e.g. 'Price Sheet SG',
    'Price Sheet - SG', 'Price Sheet_SG'), then falls back to the
    generic tab name."""
    candidates = []
    if region:
        candidates += [f"{base_name} {region}", f"{base_name}-{region}", f"{base_name}_{region}", f"{base_name} - {region}"]
    candidates.append(base_name)

    for name in candidates:
        found = find_sheet(wb, name, required=False)
        if found:
            return found
    if required:
        raise ConversionError(f"Could not find sheet '{base_name}' for region '{region}'. Available sheets: {wb.sheetnames}")
    return None


# ---------------------------------------------------------------------------
# Core conversion
# ---------------------------------------------------------------------------

def run_conversion(input_ws, price_ws, category_ws, size_chart_ws, stock_ws=None,
                   currency_code="PHP", price_col_setting="D", price_column_name=None,
                   region=None, keep_debug_writes=False,
                   progress_callback=None, custom_headings=None):
    amount_map = construct_amount_map(price_ws, price_column_name)
    category_rules = construct_category_map(category_ws)
    size_chart_map = get_template_attribute1(size_chart_ws)
    price_col_idx = parse_column_setting(input_ws, price_col_setting)

    groups = {}
    for r in range(2, input_ws.max_row + 1):
        division = val(input_ws, r, 14)
        parent_id = get_parent_key(input_ws, r, division)
        if not parent_id:
            parent_id = f"UNKNOWN_{r}"
        groups.setdefault(parent_id, []).append(r)

    main = OutputSheet()
    col = create_heading_for_target_sheet(main, custom_headings)

    # Helper function to write by canonical header/field name
    def set_by_header(row, header_name, value):
        c_idx = col.get(header_name)
        if c_idx:
            main.set_value(row, c_idx, value)

    current_out_row = 2
    total_parents = len(groups)
    p_counter = 0

    for parent_id, row_indices in groups.items():
        p_counter += 1
        if progress_callback:
            progress_callback(p_counter, total_parents)

        first_idx = row_indices[0]

        products_division = val(input_ws, first_idx, 14)
        brand = val(input_ws, first_idx, 3)
        regional_display_name = val(input_ws, first_idx, 2)
        gender = val(input_ws, first_idx, 5)
        activity_group = val(input_ws, first_idx, 8)
        article_type = val(input_ws, first_idx, 7)
        search_color_name = val(input_ws, first_idx, 13)
        long_description = val(input_ws, first_idx, 25)
        care_instruction = val(input_ws, first_idx, 43)

        item_title = get_item_title(regional_display_name, brand, gender, activity_group, article_type, search_color_name, products_division)

        # Category ID resolved from Title + Gender (most specific match wins)
        cat_id = match_category(category_rules, item_title, gender)

        size_chart_key = f"{val(input_ws, first_idx, 4)}-{gender}-{val(input_ws, first_idx, 6)}-{article_type}"
        size_chart_val = size_chart_map.get(size_chart_key)
        size_chart_attr = ("sizechart=" + s(size_chart_val[1])) if (size_chart_val and len(size_chart_val) > 1) else ""

        def resolve_price(row_idx):
            """The user-selected Price Tracker column always wins; the
            Input sheet's fallback column is only used if the SKU has no
            entry (or a blank value) in the Price Tracker."""
            row_sku = s(val(input_ws, row_idx, 16)).strip()
            amt = amount_map.get(row_sku) if row_sku else None
            tracker_price = amt.get("price", "") if amt else ""
            if tracker_price not in ("", None):
                return tracker_price
            return val(input_ws, row_idx, price_col_idx)

        parent_rrp = resolve_price(first_idx)
        short_desc = build_short_description(input_ws, first_idx)

        desc_attr = f"description={replace_spl_character(long_description)}" if long_description else ""
        care_attr = f"care={replace_spl_character(care_instruction)}" if care_instruction else ""

        # -------------------------------------------------------------------
        # 1. INSERT PARENT ROW
        # -------------------------------------------------------------------
        set_by_header(current_out_row, "SKU", parent_id)
        set_by_header(current_out_row, "customSKU", parent_id)
        set_by_header(current_out_row, "itemTitle", replace_spl_character(item_title))
        set_by_header(current_out_row, "noOfVariants", len(row_indices))
        set_by_header(current_out_row, "variation1", "color_family")
        set_by_header(current_out_row, "variation2", "size")
        set_by_header(current_out_row, "shortDescription", replace_spl_character(short_desc))
        set_by_header(current_out_row, "itemAmount", parent_rrp)
        set_by_header(current_out_row, "currencyCode", currency_code)
        set_by_header(current_out_row, "noOfItem", 0)  # Quantity always 0
        set_by_header(current_out_row, "categoryID", cat_id)
        set_by_header(current_out_row, "brand", brand)
        set_by_header(current_out_row, "packageContent", f"1 X {replace_spl_character(item_title)}")

        if size_chart_attr:
            set_by_header(current_out_row, "templateAttribute1", size_chart_attr)
        if desc_attr:
            set_by_header(current_out_row, "templateAttribute2", desc_attr)
        if care_attr:
            set_by_header(current_out_row, "templateAttribute4", care_attr)

        current_out_row += 1

        # -------------------------------------------------------------------
        # 2. SORT AND INSERT VARIANT ROWS
        # -------------------------------------------------------------------
        sorted_row_indices = sorted(
            row_indices,
            key=lambda idx: size_sort_key(get_variation2_size(input_ws, idx))
        )

        for r_idx in sorted_row_indices:
            custom_sku = s(val(input_ws, r_idx, 16)).strip()
            variant_rrp = resolve_price(r_idx)
            color_name = val(input_ws, r_idx, 9)              # Color Name from Input Sheet
            size_uk_val = get_variation2_size(input_ws, r_idx)  # Size UK from Input Sheet

            amt_val = amount_map.get(custom_sku) if custom_sku else None
            sale_price = amt_val.get("sale_price", "") if amt_val else ""

            set_by_header(current_out_row, "SKU", parent_id)
            set_by_header(current_out_row, "customSKU", custom_sku)
            set_by_header(current_out_row, "itemTitle", replace_spl_character(item_title))
            set_by_header(current_out_row, "variation1", color_name)
            set_by_header(current_out_row, "variation2", size_uk_val)
            set_by_header(current_out_row, "shortDescription", replace_spl_character(short_desc))
            set_by_header(current_out_row, "salePrice", sale_price)
            set_by_header(current_out_row, "itemAmount", variant_rrp)
            set_by_header(current_out_row, "currencyCode", currency_code)
            set_by_header(current_out_row, "noOfItem", 0)  # Quantity always 0
            set_by_header(current_out_row, "categoryID", cat_id)
            set_by_header(current_out_row, "brand", brand)
            set_by_header(current_out_row, "packageContent", f"1 X {replace_spl_character(item_title)}")

            if size_chart_attr:
                set_by_header(current_out_row, "templateAttribute1", size_chart_attr)
            if desc_attr:
                set_by_header(current_out_row, "templateAttribute2", desc_attr)
            if care_attr:
                set_by_header(current_out_row, "templateAttribute4", care_attr)

            current_out_row += 1

    return main


def build_output_workbook(input_bytes, price_bytes=None, category_bytes=None,
                          size_chart_bytes=None, sample_output_bytes=None,
                          region="PH", currency_code=None, price_col_setting="D",
                          price_column_name=None,
                          keep_debug_writes=False, progress_callback=None):
    try:
        wb = load_workbook(io.BytesIO(input_bytes), data_only=True)
    except Exception as e:
        raise ConversionError(f"Could not open main input workbook: {e}")

    input_ws = find_sheet(wb, "Input")

    resolved_currency = currency_code or REGION_CURRENCY.get(region, "PHP")

    def get_target_ws(file_bytes, sheet_name):
        if file_bytes:
            temp_wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            return temp_wb.active
        return find_regional_sheet(wb, sheet_name, region, required=False)

    price_ws = get_target_ws(price_bytes, "Price Sheet")
    category_ws = get_target_ws(category_bytes, "Category sheet")
    size_chart_ws = get_target_ws(size_chart_bytes, "Size chart")
    stock_ws = find_sheet(wb, "Stock sheet", required=False)

    if not price_ws:
        raise ConversionError("Price Sheet missing from main file and no standalone upload provided.")
    if not category_ws:
        raise ConversionError("Category Sheet missing from main file and no standalone upload provided.")
    if not size_chart_ws:
        raise ConversionError("Size Chart Sheet missing from main file and no standalone upload provided.")

    custom_headings = None
    if sample_output_bytes:
        sample_wb = load_workbook(io.BytesIO(sample_output_bytes), data_only=True)
        sample_ws = sample_wb.active
        custom_headings = [val(sample_ws, 1, c) for c in range(1, sample_ws.max_column + 1) if val(sample_ws, 1, c) != ""]

    output = run_conversion(
        input_ws, price_ws, category_ws, size_chart_ws,
        stock_ws=stock_ws, currency_code=resolved_currency,
        price_col_setting=price_col_setting,
        price_column_name=price_column_name,
        region=region,
        keep_debug_writes=keep_debug_writes,
        progress_callback=progress_callback,
        custom_headings=custom_headings
    )

    out_wb = output.to_workbook()
    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.read()
